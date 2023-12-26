import csv
import inspect
import logging

import softest
from openpyxl import workbook, load_workbook

class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for stop in list:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEquals,stop.text,value)
            if stop.text == value:
                print("assert pass")
            else:
                print("assert failed")
        self.assert_all()


    def  custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        # create console handler and set level to debug

        fh = logging.FileHandler("automation.log",mode="w")
        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s',datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to ch
        fh.setFormatter(formatter)
        # add ch to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(self,filename,sheet):

        wb = load_workbook(filename=filename)
        datalist=[]
        sh = wb[sheet]
        row_count = sh.max_row
        column_count = sh.max_column

        for i in range(1, row_count + 1):
            row=[]
            for j in range(1, column_count + 1):
                print(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        # create an empty list
        csvlist=[]
#         open csv file
        csvdata = open(filename,"r")

        # create csv reader
        reader = csv.reader(csvdata)
        # skip header
        next(reader)

        # add csv rows to list
        for rows in reader:
            csvlist.append(rows)
        return csvlist

